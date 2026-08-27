import uuid
from datetime import date, datetime, timezone
from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy import select, func
from sqlalchemy.orm import aliased
from sqlalchemy.ext.asyncio import AsyncSession
from pydantic import BaseModel
from typing import Optional
from app.core.database import get_db
from app.api.deps import require_staff, require_owner, get_current_user
from app.models.user import User, UserRole
from app.models.finance import (
    FinancialAccount, AccountType, Expense, ExpenseEdit, ExpenseCategoryDef, CATEGORY_LABEL,
    AccountTransfer,
)
from app.models.payment import Payment, PaymentStatus
from app.schemas.common import Page
from app.schemas.finance import (
    AccountCreate, AccountUpdate, AccountResponse,
    ExpenseCreate, ExpenseUpdate, ExpenseRow, ExpenseEditRow,
    ExpenseCategoryRow, ExpenseCategoryCreate, ExpenseCategoryUpdate,
    FinanceReport, CategoryAmount, LedgerEntry, LedgerResponse,
    TransferCreate, TransferRow,
)
from app.services.finance import account_balance


async def _category_labels(db: AsyncSession) -> dict:
    """Peta {key: label} dari tabel kategori (fallback ke CATEGORY_LABEL bawaan / key mentah)."""
    rows = (await db.execute(select(ExpenseCategoryDef.key, ExpenseCategoryDef.label))).all()
    m = {k: lbl for k, lbl in rows}
    return m


def _label_for(m: dict, key: str) -> str:
    return m.get(key) or CATEGORY_LABEL.get(key, key)


def _slugify(text: str) -> str:
    import re
    s = re.sub(r"[^a-z0-9]+", "-", (text or "").strip().lower()).strip("-")
    return s or "kategori"

router = APIRouter()


def _fmt_rp(v) -> str:
    return f"Rp{int(round(float(v or 0))):,}".replace(",", ".")


def _fmt_date(d) -> str:
    return f"{d.day:02d}/{d.month:02d}/{d.year}" if d else "—"


# ── util Excel ──
_COPPER = "8A5140"


def _new_workbook(sheet_title: str):
    """Buat workbook + gaya standar (title/head/bold/normal/muted/money/fill)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    styles = {
        "title": Font(bold=True, size=14, color=_COPPER),
        "head": Font(bold=True, color=_COPPER),
        "bold": Font(bold=True),
        "normal": Font(),
        "muted": Font(size=10, color="888888"),
        "money": "#,##0",
        "fill": PatternFill("solid", fgColor="F0E0D6"),
    }
    return wb, ws, styles


def _xlsx_response(wb, fname: str):
    import io
    from fastapi.responses import StreamingResponse
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


class TransferAccount(BaseModel):
    name: str
    bank_name: Optional[str] = None
    account_number: Optional[str] = None

    class Config:
        from_attributes = True


@router.get("/transfer-info", response_model=list[TransferAccount])
async def transfer_info(db: AsyncSession = Depends(get_db), _: User = Depends(get_current_user)):
    """Rekening bank aktif untuk instruksi transfer (dilihat member saat bayar)."""
    rows = (
        await db.execute(
            select(FinancialAccount).where(
                FinancialAccount.is_active.is_(True), FinancialAccount.type == AccountType.BANK
            ).order_by(FinancialAccount.created_at)
        )
    ).scalars().all()
    return rows


# ─────────────── AKUN KAS/BANK ───────────────
async def _account_out(db: AsyncSession, acc: FinancialAccount) -> AccountResponse:
    r = AccountResponse.model_validate(acc)
    r.balance = await account_balance(db, acc)
    return r


@router.get("/accounts", response_model=list[AccountResponse])
async def list_accounts(
    include_inactive: bool = Query(False),
    db: AsyncSession = Depends(get_db), user: User = Depends(require_staff),
):
    stmt = select(FinancialAccount)
    if not include_inactive:
        stmt = stmt.where(FinancialAccount.is_active.is_(True))
    rows = (await db.execute(stmt.order_by(FinancialAccount.type, FinancialAccount.created_at))).scalars().all()
    is_owner = user.role == UserRole.OWNER
    out = []
    for a in rows:
        r = await _account_out(db, a)
        # Non-owner hanya boleh lihat saldo KAS studio; saldo rekening bank disembunyikan.
        if not is_owner and a.type != AccountType.CASH:
            r.balance = None
        out.append(r)
    return out


@router.post("/accounts", response_model=AccountResponse, status_code=201)
async def create_account(payload: AccountCreate, db: AsyncSession = Depends(get_db), _: User = Depends(require_staff)):
    acc = FinancialAccount(**payload.model_dump())
    db.add(acc)
    await db.flush()
    await db.refresh(acc)
    return await _account_out(db, acc)


@router.patch("/accounts/{account_id}", response_model=AccountResponse)
async def update_account(account_id: uuid.UUID, payload: AccountUpdate, db: AsyncSession = Depends(get_db), _: User = Depends(require_staff)):
    acc = (await db.execute(select(FinancialAccount).where(FinancialAccount.id == account_id))).scalar_one_or_none()
    if not acc:
        raise HTTPException(404, "Akun tidak ditemukan")
    for k, v in payload.model_dump(exclude_unset=True).items():
        setattr(acc, k, v)
    await db.flush()
    await db.refresh(acc)
    return await _account_out(db, acc)


@router.delete("/accounts/{account_id}", status_code=204)
async def delete_account(account_id: uuid.UUID, db: AsyncSession = Depends(get_db), _: User = Depends(require_staff)):
    """Non-aktifkan akun (data pengeluaran/income tetap tersimpan)."""
    acc = (await db.execute(select(FinancialAccount).where(FinancialAccount.id == account_id))).scalar_one_or_none()
    if not acc:
        raise HTTPException(404, "Akun tidak ditemukan")
    acc.is_active = False
    return None


# ── Transfer antar akun ──
async def _transfer_rows(db: AsyncSession, date_from=None, date_to=None) -> list[TransferRow]:
    FromAcc = aliased(FinancialAccount)
    ToAcc = aliased(FinancialAccount)
    stmt = (
        select(AccountTransfer, FromAcc.name, ToAcc.name)
        .outerjoin(FromAcc, AccountTransfer.from_account_id == FromAcc.id)
        .outerjoin(ToAcc, AccountTransfer.to_account_id == ToAcc.id)
        .order_by(AccountTransfer.transfer_date.desc(), AccountTransfer.created_at.desc())
    )
    if date_from is not None:
        stmt = stmt.where(AccountTransfer.transfer_date >= date_from)
    if date_to is not None:
        stmt = stmt.where(AccountTransfer.transfer_date <= date_to)
    rows = (await db.execute(stmt)).all()
    return [
        TransferRow(
            id=t.id, transfer_date=t.transfer_date, amount=float(t.amount or 0),
            from_account_id=t.from_account_id, to_account_id=t.to_account_id,
            from_account_name=fn, to_account_name=tn, description=t.description, created_at=t.created_at,
        )
        for t, fn, tn in rows
    ]


@router.get("/transfers", response_model=list[TransferRow])
async def list_transfers(
    date_from: date | None = Query(None, alias="from"),
    date_to: date | None = Query(None, alias="to"),
    db: AsyncSession = Depends(get_db), _: User = Depends(require_staff),
):
    return await _transfer_rows(db, date_from, date_to)


@router.post("/transfers", response_model=TransferRow, status_code=201)
async def create_transfer(payload: TransferCreate, db: AsyncSession = Depends(get_db), staff: User = Depends(require_staff)):
    if payload.from_account_id == payload.to_account_id:
        raise HTTPException(400, "Akun asal dan tujuan tidak boleh sama")
    ids = {payload.from_account_id, payload.to_account_id}
    accs = (await db.execute(select(FinancialAccount).where(FinancialAccount.id.in_(ids)))).scalars().all()
    if len(accs) != 2:
        raise HTTPException(400, "Akun asal/tujuan tidak ditemukan")
    t = AccountTransfer(
        transfer_date=payload.transfer_date, from_account_id=payload.from_account_id,
        to_account_id=payload.to_account_id, amount=payload.amount,
        description=payload.description, recorded_by_id=staff.id,
    )
    db.add(t)
    await db.flush()
    rows = await _transfer_rows(db)
    return next(r for r in rows if r.id == t.id)


@router.delete("/transfers/{transfer_id}", status_code=204)
async def delete_transfer(transfer_id: uuid.UUID, db: AsyncSession = Depends(get_db), _: User = Depends(require_staff)):
    t = (await db.execute(select(AccountTransfer).where(AccountTransfer.id == transfer_id))).scalar_one_or_none()
    if not t:
        raise HTTPException(404, "Transfer tidak ditemukan")
    await db.delete(t)
    return None


async def _build_ledger(db: AsyncSession, account_id: uuid.UUID, date_from, date_to) -> LedgerResponse:
    """Buku besar akun: mutasi masuk (pembayaran lunas) & keluar (pengeluaran) + saldo berjalan.
    Dihitung dari SELURUH riwayat sejak saldo awal agar rekonsiliasi dgn saldo akun."""
    acc = (await db.execute(select(FinancialAccount).where(FinancialAccount.id == account_id))).scalar_one_or_none()
    if not acc:
        raise HTTPException(404, "Akun tidak ditemukan")

    # Uang masuk = pembayaran LUNAS yang ter-atribusi ke akun ini
    inc_rows = (
        await db.execute(
            select(
                Payment.amount, func.coalesce(Payment.paid_at, Payment.created_at), Payment.note, User.full_name
            ).outerjoin(User, Payment.member_id == User.id)
            .where(Payment.account_id == account_id, Payment.status == PaymentStatus.PAID)
        )
    ).all()
    # Uang keluar = pengeluaran dari akun ini
    exp_rows = (
        await db.execute(
            select(Expense.amount, Expense.expense_date, Expense.category, Expense.description)
            .where(Expense.account_id == account_id)
        )
    ).all()

    # Transfer masuk (akun ini = tujuan) & keluar (akun ini = asal)
    FromAcc = aliased(FinancialAccount)
    ToAcc = aliased(FinancialAccount)
    tr_in_rows = (
        await db.execute(
            select(AccountTransfer.amount, AccountTransfer.transfer_date, AccountTransfer.description, FromAcc.name)
            .outerjoin(FromAcc, AccountTransfer.from_account_id == FromAcc.id)
            .where(AccountTransfer.to_account_id == account_id)
        )
    ).all()
    tr_out_rows = (
        await db.execute(
            select(AccountTransfer.amount, AccountTransfer.transfer_date, AccountTransfer.description, ToAcc.name)
            .outerjoin(ToAcc, AccountTransfer.to_account_id == ToAcc.id)
            .where(AccountTransfer.from_account_id == account_id)
        )
    ).all()

    cat_labels = await _category_labels(db)
    events = []  # (sort_key_datetime, date, kind, description, amount)
    for amount, ts, note, member_name in inc_rows:
        d = ts.date()
        label = note or "Pembayaran"
        if member_name:
            label = f"{label} — {member_name}"
        events.append((ts.replace(tzinfo=None), d, "in", label, float(amount or 0)))
    for amount, edate, category, description in exp_rows:
        label = _label_for(cat_labels, category)
        if description:
            label = f"{label} — {description}"
        events.append((datetime(edate.year, edate.month, edate.day), edate, "out", label, float(amount or 0)))
    for amount, tdate, description, from_name in tr_in_rows:
        label = f"Transfer dari {from_name or '—'}"
        if description:
            label = f"{label} — {description}"
        events.append((datetime(tdate.year, tdate.month, tdate.day), tdate, "in", label, float(amount or 0)))
    for amount, tdate, description, to_name in tr_out_rows:
        label = f"Transfer ke {to_name or '—'}"
        if description:
            label = f"{label} — {description}"
        events.append((datetime(tdate.year, tdate.month, tdate.day), tdate, "out", label, float(amount or 0)))

    events.sort(key=lambda e: e[0])

    opening = float(acc.opening_balance or 0)
    running = opening
    starting_balance = opening
    entries: list[LedgerEntry] = []
    total_in = 0.0
    total_out = 0.0
    for _sk, d, kind, label, amount in events:
        if kind == "in":
            running += amount
        else:
            running -= amount
        # sebelum periode → hanya update saldo awal periode, tidak ditampilkan
        if date_from and d < date_from:
            starting_balance = running
            continue
        if date_to and d > date_to:
            continue
        if kind == "in":
            total_in += amount
        else:
            total_out += amount
        entries.append(LedgerEntry(date=d, kind=kind, description=label, amount=amount, balance=running))

    ending_balance = entries[-1].balance if entries else starting_balance
    return LedgerResponse(
        account_id=acc.id, account_name=acc.name, account_type=acc.type,
        opening_balance=opening, starting_balance=starting_balance,
        total_in=total_in, total_out=total_out, ending_balance=ending_balance,
        entries=entries,
    )


@router.get("/accounts/{account_id}/ledger", response_model=LedgerResponse)
async def account_ledger(
    account_id: uuid.UUID,
    date_from: date | None = Query(None, alias="from"),
    date_to: date | None = Query(None, alias="to"),
    db: AsyncSession = Depends(get_db), _: User = Depends(require_owner),
):
    return await _build_ledger(db, account_id, date_from, date_to)


@router.get("/accounts/{account_id}/ledger.xlsx")
async def account_ledger_xlsx(
    account_id: uuid.UUID,
    date_from: date | None = Query(None, alias="from"),
    date_to: date | None = Query(None, alias="to"),
    db: AsyncSession = Depends(get_db), _: User = Depends(require_owner),
):
    """Ekspor buku besar akun ke Excel (.xlsx)."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from fastapi.responses import StreamingResponse

    data = await _build_ledger(db, account_id, date_from, date_to)

    wb = Workbook()
    ws = wb.active
    ws.title = "Buku Besar"

    copper = "8A5140"
    head_fill = PatternFill("solid", fgColor="F0E0D6")
    money = "#,##0"
    thin = Side(style="thin", color="E5D5CB")
    border = Border(bottom=thin)

    ws.merge_cells("A1:E1")
    c = ws["A1"]; c.value = f"Buku Besar — {data.account_name}"
    c.font = Font(bold=True, size=14, color=copper)
    ws.merge_cells("A2:E2")
    periode = f"Periode {date_from.strftime('%d/%m/%Y') if date_from else 'awal'} – {date_to.strftime('%d/%m/%Y') if date_to else 'kini'}"
    ws["A2"].value = periode
    ws["A2"].font = Font(size=10, color="888888")

    headers = ["Tanggal", "Keterangan", "Masuk", "Keluar", "Saldo"]
    hr = 4
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=hr, column=i, value=h)
        cell.font = Font(bold=True, color=copper)
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="right" if i >= 3 else "left")

    r = hr + 1
    # baris saldo awal periode
    ws.cell(row=r, column=1, value=(date_from.strftime("%d/%m/%Y") if date_from else ""))
    ws.cell(row=r, column=2, value="Saldo awal periode").font = Font(italic=True)
    sc = ws.cell(row=r, column=5, value=float(data.starting_balance)); sc.number_format = money
    for col in range(1, 6):
        ws.cell(row=r, column=col).fill = head_fill
    r += 1

    for e in data.entries:
        ws.cell(row=r, column=1, value=e.date.strftime("%d/%m/%Y"))
        ws.cell(row=r, column=2, value=e.description)
        if e.kind == "in":
            ic = ws.cell(row=r, column=3, value=float(e.amount)); ic.number_format = money
        else:
            oc = ws.cell(row=r, column=4, value=float(e.amount)); oc.number_format = money
        bc = ws.cell(row=r, column=5, value=float(e.balance)); bc.number_format = money
        for col in range(1, 6):
            ws.cell(row=r, column=col).border = border
        r += 1

    # ringkasan
    r += 1
    ws.cell(row=r, column=2, value="Total Masuk").font = Font(bold=True)
    tc = ws.cell(row=r, column=3, value=float(data.total_in)); tc.number_format = money; tc.font = Font(bold=True)
    r += 1
    ws.cell(row=r, column=2, value="Total Keluar").font = Font(bold=True)
    oc = ws.cell(row=r, column=4, value=float(data.total_out)); oc.number_format = money; oc.font = Font(bold=True)
    r += 1
    ws.cell(row=r, column=2, value="Saldo Akhir").font = Font(bold=True, color=copper)
    ec = ws.cell(row=r, column=5, value=float(data.ending_balance)); ec.number_format = money; ec.font = Font(bold=True, color=copper)

    widths = [14, 46, 16, 16, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe = "".join(ch for ch in data.account_name if ch.isalnum() or ch in " -_").strip().replace(" ", "_")
    fname = f"BukuBesar_{safe or 'akun'}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ─────────────── KATEGORI PENGELUARAN ───────────────
@router.get("/expense-categories", response_model=list[ExpenseCategoryRow])
async def list_expense_categories(
    include_inactive: bool = Query(False),
    db: AsyncSession = Depends(get_db), _: User = Depends(require_staff),
):
    stmt = select(ExpenseCategoryDef)
    if not include_inactive:
        stmt = stmt.where(ExpenseCategoryDef.is_active.is_(True))
    rows = (await db.execute(stmt.order_by(ExpenseCategoryDef.sort_order, ExpenseCategoryDef.label))).scalars().all()
    return rows


@router.post("/expense-categories", response_model=ExpenseCategoryRow, status_code=201)
async def create_expense_category(payload: ExpenseCategoryCreate, db: AsyncSession = Depends(get_db), _: User = Depends(require_staff)):
    label = payload.label.strip()
    base = _slugify(label)
    key = base
    i = 2
    while (await db.execute(select(ExpenseCategoryDef.id).where(ExpenseCategoryDef.key == key))).scalar_one_or_none():
        key = f"{base}-{i}"; i += 1
    cat = ExpenseCategoryDef(key=key, label=label, is_active=True, is_builtin=False, sort_order=500)
    db.add(cat)
    await db.flush()
    await db.refresh(cat)
    return cat


@router.patch("/expense-categories/{cat_id}", response_model=ExpenseCategoryRow)
async def update_expense_category(cat_id: uuid.UUID, payload: ExpenseCategoryUpdate, db: AsyncSession = Depends(get_db), _: User = Depends(require_staff)):
    cat = (await db.execute(select(ExpenseCategoryDef).where(ExpenseCategoryDef.id == cat_id))).scalar_one_or_none()
    if not cat:
        raise HTTPException(404, "Kategori tidak ditemukan")
    data = payload.model_dump(exclude_unset=True)
    if "label" in data and data["label"]:
        cat.label = data["label"].strip()
    if "is_active" in data and data["is_active"] is not None:
        if cat.is_builtin and data["is_active"] is False:
            raise HTTPException(400, "Kategori bawaan tidak bisa dinonaktifkan")
        cat.is_active = data["is_active"]
    await db.flush()
    await db.refresh(cat)
    return cat


@router.delete("/expense-categories/{cat_id}", status_code=204)
async def delete_expense_category(cat_id: uuid.UUID, db: AsyncSession = Depends(get_db), _: User = Depends(require_staff)):
    cat = (await db.execute(select(ExpenseCategoryDef).where(ExpenseCategoryDef.id == cat_id))).scalar_one_or_none()
    if not cat:
        raise HTTPException(404, "Kategori tidak ditemukan")
    if cat.is_builtin:
        raise HTTPException(400, "Kategori bawaan tidak bisa dihapus (bisa dinonaktifkan lewat tambahan saja)")
    used = (await db.execute(select(func.count()).select_from(Expense).where(Expense.category == cat.key))).scalar_one()
    if used:
        raise HTTPException(400, f"Kategori dipakai {used} pengeluaran — nonaktifkan saja, jangan hapus")
    await db.delete(cat)
    return None


# ─────────────── PENGELUARAN ───────────────
@router.get("/expenses", response_model=Page[ExpenseRow])
async def list_expenses(
    date_from: date | None = Query(None, alias="from"),
    date_to: date | None = Query(None, alias="to"),
    category: str | None = Query(None),
    account_id: uuid.UUID | None = Query(None),
    limit: int = Query(20, le=200),
    offset: int = Query(0, ge=0),
    db: AsyncSession = Depends(get_db), _: User = Depends(require_staff),
):
    edit_count_sq = (
        select(func.count(ExpenseEdit.id)).where(ExpenseEdit.expense_id == Expense.id).scalar_subquery()
    )
    stmt = select(Expense, FinancialAccount.name, edit_count_sq).outerjoin(FinancialAccount, Expense.account_id == FinancialAccount.id)
    conds = []
    if date_from:
        conds.append(Expense.expense_date >= date_from)
    if date_to:
        conds.append(Expense.expense_date <= date_to)
    if category:
        conds.append(Expense.category == category)
    if account_id:
        conds.append(Expense.account_id == account_id)
    for c in conds:
        stmt = stmt.where(c)

    count_stmt = select(func.count()).select_from(Expense)
    for c in conds:
        count_stmt = count_stmt.where(c)
    total = (await db.execute(count_stmt)).scalar_one()

    rows = (await db.execute(stmt.order_by(Expense.expense_date.desc(), Expense.created_at.desc()).limit(limit).offset(offset))).all()
    items = []
    for exp, acc_name, ecount in rows:
        row = ExpenseRow.model_validate(exp)
        row.account_name = acc_name
        row.edit_count = ecount or 0
        items.append(row)
    return Page(items=items, total=total)


@router.get("/expenses.xlsx")
async def expenses_xlsx(
    date_from: date | None = Query(None, alias="from"),
    date_to: date | None = Query(None, alias="to"),
    category: str | None = Query(None),
    account_id: uuid.UUID | None = Query(None),
    db: AsyncSession = Depends(get_db), _: User = Depends(require_staff),
):
    """Ekspor daftar pengeluaran (sesuai filter) ke Excel."""
    stmt = select(Expense, FinancialAccount.name).outerjoin(FinancialAccount, Expense.account_id == FinancialAccount.id)
    conds = []
    if date_from:
        conds.append(Expense.expense_date >= date_from)
    if date_to:
        conds.append(Expense.expense_date <= date_to)
    if category:
        conds.append(Expense.category == category)
    if account_id:
        conds.append(Expense.account_id == account_id)
    for c in conds:
        stmt = stmt.where(c)
    rows = (await db.execute(stmt.order_by(Expense.expense_date.desc(), Expense.created_at.desc()))).all()

    wb, ws, S = _new_workbook("Pengeluaran")
    ws.merge_cells("A1:E1"); ws["A1"].value = "Daftar Pengeluaran"; ws["A1"].font = S["title"]
    prd = f"Periode {date_from.strftime('%d/%m/%Y') if date_from else 'awal'} – {date_to.strftime('%d/%m/%Y') if date_to else 'kini'}"
    ws.merge_cells("A2:E2"); ws["A2"].value = prd; ws["A2"].font = S["muted"]

    hr = 4
    for i, h in enumerate(["Tanggal", "Kategori", "Keterangan", "Akun", "Jumlah"], start=1):
        cell = ws.cell(row=hr, column=i, value=h)
        cell.font = S["head"]; cell.fill = S["fill"]
    cat_labels = await _category_labels(db)
    r = hr + 1
    total = 0.0
    for exp, acc_name in rows:
        ws.cell(row=r, column=1, value=exp.expense_date.strftime("%d/%m/%Y"))
        ws.cell(row=r, column=2, value=_label_for(cat_labels, exp.category))
        ws.cell(row=r, column=3, value=exp.description or "")
        ws.cell(row=r, column=4, value=acc_name or "")
        amt = ws.cell(row=r, column=5, value=float(exp.amount or 0)); amt.number_format = S["money"]
        total += float(exp.amount or 0)
        r += 1
    ws.cell(row=r, column=4, value="Total").font = S["bold"]
    tc = ws.cell(row=r, column=5, value=total); tc.number_format = S["money"]; tc.font = S["bold"]

    for col, w in zip("ABCDE", [14, 22, 40, 22, 18]):
        ws.column_dimensions[col].width = w
    from app.services.booking import today_local
    return _xlsx_response(wb, f"Pengeluaran_{(date_from or today_local()).strftime('%Y%m%d')}.xlsx")


async def _validate_category(db: AsyncSession, key: str) -> None:
    cat = (await db.execute(select(ExpenseCategoryDef).where(ExpenseCategoryDef.key == key))).scalar_one_or_none()
    if not cat or not cat.is_active:
        raise HTTPException(400, "Kategori tidak valid / nonaktif")


@router.post("/expenses", response_model=ExpenseRow, status_code=201)
async def create_expense(payload: ExpenseCreate, db: AsyncSession = Depends(get_db), staff: User = Depends(require_staff)):
    acc = (await db.execute(select(FinancialAccount).where(FinancialAccount.id == payload.account_id))).scalar_one_or_none()
    if not acc:
        raise HTTPException(400, "Akun tidak valid")
    await _validate_category(db, payload.category)
    exp = Expense(**payload.model_dump(), recorded_by_id=staff.id)
    db.add(exp)
    await db.flush()
    await db.refresh(exp)
    row = ExpenseRow.model_validate(exp)
    row.account_name = acc.name
    return row


async def _account_name(db: AsyncSession, account_id) -> str:
    if not account_id:
        return "—"
    name = (await db.execute(select(FinancialAccount.name).where(FinancialAccount.id == account_id))).scalar_one_or_none()
    return name or "—"


@router.patch("/expenses/{expense_id}", response_model=ExpenseRow)
async def update_expense(expense_id: uuid.UUID, payload: ExpenseUpdate, db: AsyncSession = Depends(get_db), staff: User = Depends(require_staff)):
    exp = (await db.execute(select(Expense).where(Expense.id == expense_id))).scalar_one_or_none()
    if not exp:
        raise HTTPException(404, "Pengeluaran tidak ditemukan")

    data = payload.model_dump(exclude_unset=True)
    if "category" in data and data["category"] is not None and data["category"] != exp.category:
        await _validate_category(db, data["category"])
    cat_labels = await _category_labels(db)
    changes: list[str] = []
    for field, new_val in data.items():
        old_val = getattr(exp, field)
        if field == "amount":
            if float(old_val) == float(new_val):
                continue
            changes.append(f"Jumlah: {_fmt_rp(old_val)} → {_fmt_rp(new_val)}")
        elif field == "category":
            if old_val == new_val:
                continue
            changes.append(f"Kategori: {_label_for(cat_labels, old_val)} → {_label_for(cat_labels, new_val)}")
        elif field == "expense_date":
            if old_val == new_val:
                continue
            changes.append(f"Tanggal: {_fmt_date(old_val)} → {_fmt_date(new_val)}")
        elif field == "account_id":
            if old_val == new_val:
                continue
            changes.append(f"Akun: {await _account_name(db, old_val)} → {await _account_name(db, new_val)}")
        elif field == "description":
            if (old_val or "") == (new_val or ""):
                continue
            changes.append(f"Keterangan: {old_val or '—'} → {new_val or '—'}")
        setattr(exp, field, new_val)

    if changes:
        db.add(ExpenseEdit(
            expense_id=exp.id, edited_by_id=staff.id, edited_by_name=staff.full_name,
            summary="; ".join(changes),
        ))

    await db.flush()
    await db.refresh(exp)
    ecount = (await db.execute(select(func.count(ExpenseEdit.id)).where(ExpenseEdit.expense_id == exp.id))).scalar_one()
    row = ExpenseRow.model_validate(exp)
    row.account_name = await _account_name(db, exp.account_id) if exp.account_id else None
    row.edit_count = ecount or 0
    return row


@router.get("/expenses/{expense_id}/history", response_model=list[ExpenseEditRow])
async def expense_history(expense_id: uuid.UUID, db: AsyncSession = Depends(get_db), _: User = Depends(require_staff)):
    rows = (
        await db.execute(
            select(ExpenseEdit).where(ExpenseEdit.expense_id == expense_id).order_by(ExpenseEdit.created_at.desc())
        )
    ).scalars().all()
    return rows


@router.delete("/expenses/{expense_id}", status_code=204)
async def delete_expense(expense_id: uuid.UUID, db: AsyncSession = Depends(get_db), _: User = Depends(require_staff)):
    exp = (await db.execute(select(Expense).where(Expense.id == expense_id))).scalar_one_or_none()
    if not exp:
        raise HTTPException(404, "Pengeluaran tidak ditemukan")
    await db.delete(exp)
    return None


# ─────────────── LAPORAN KEUANGAN ───────────────
async def _compute_report(db: AsyncSession, date_from: date, date_to: date) -> FinanceReport:
    # Income = pembayaran LUNAS pada rentang (pakai created_at tanggalnya)
    income = (
        await db.execute(
            select(func.coalesce(func.sum(Payment.amount), 0)).where(
                Payment.status == PaymentStatus.PAID,
                func.date(Payment.created_at) >= date_from,
                func.date(Payment.created_at) <= date_to,
            )
        )
    ).scalar_one()

    in_range = (Expense.expense_date >= date_from) & (Expense.expense_date <= date_to)
    expense = (await db.execute(select(func.coalesce(func.sum(Expense.amount), 0)).where(in_range))).scalar_one()

    cat_rows = (
        await db.execute(
            select(Expense.category, func.coalesce(func.sum(Expense.amount), 0)).where(in_range).group_by(Expense.category)
        )
    ).all()
    cat_labels = await _category_labels(db)
    by_cat = [CategoryAmount(category=c, label=_label_for(cat_labels, c), amount=float(a or 0)) for c, a in cat_rows]

    accounts = (
        await db.execute(select(FinancialAccount).where(FinancialAccount.is_active.is_(True)).order_by(FinancialAccount.type, FinancialAccount.created_at))
    ).scalars().all()
    acc_out = [await _account_out(db, a) for a in accounts]
    transfers = await _transfer_rows(db, date_from, date_to)

    return FinanceReport(
        date_from=date_from, date_to=date_to,
        income=float(income or 0), expense=float(expense or 0), net=float(income or 0) - float(expense or 0),
        expense_by_category=by_cat, accounts=acc_out, transfers=transfers,
    )


@router.get("/report", response_model=FinanceReport)
async def finance_report(
    date_from: date = Query(..., alias="from"),
    date_to: date = Query(..., alias="to"),
    db: AsyncSession = Depends(get_db), _: User = Depends(require_owner),
):
    return await _compute_report(db, date_from, date_to)


@router.get("/report.xlsx")
async def report_xlsx(
    date_from: date = Query(..., alias="from"),
    date_to: date = Query(..., alias="to"),
    db: AsyncSession = Depends(get_db), _: User = Depends(require_owner),
):
    """Ekspor Laba/Rugi ke Excel."""
    data = await _compute_report(db, date_from, date_to)
    wb, ws, S = _new_workbook("Laba Rugi")
    ws.merge_cells("A1:B1"); ws["A1"].value = "Laporan Laba / Rugi"; ws["A1"].font = S["title"]
    ws.merge_cells("A2:B2"); ws["A2"].value = f"Periode {date_from.strftime('%d/%m/%Y')} – {date_to.strftime('%d/%m/%Y')}"; ws["A2"].font = S["muted"]

    r = 4
    def section(title):
        nonlocal r
        ws.cell(row=r, column=1, value=title).font = S["head"]; r += 1
    def line(label, amount, bold=False):
        nonlocal r
        ws.cell(row=r, column=1, value=label).font = S["bold"] if bold else S["normal"]
        c = ws.cell(row=r, column=2, value=float(amount)); c.number_format = S["money"]; c.font = S["bold"] if bold else S["normal"]
        r += 1

    section("PENDAPATAN")
    line("Pendapatan member & kelas", data.income)
    line("Total Pendapatan", data.income, bold=True)
    r += 1
    section("PENGELUARAN OPERASIONAL")
    if data.expense_by_category:
        for ca in sorted(data.expense_by_category, key=lambda x: x.amount, reverse=True):
            line(ca.label or ca.category, ca.amount)
    else:
        ws.cell(row=r, column=1, value="Tidak ada pengeluaran").font = S["muted"]; r += 1
    line("Total Pengeluaran", data.expense, bold=True)
    r += 1
    line("LABA BERSIH" if data.net >= 0 else "RUGI BERSIH", data.net, bold=True)
    r += 2
    section("SALDO AKUN (saat cetak)")
    for a in data.accounts:
        line(f"{a.name}{f' ({a.bank_name})' if a.bank_name else ''}", a.balance)
    line("Total Saldo", sum(a.balance for a in data.accounts), bold=True)

    if data.transfers:
        r += 2
        section("TRANSFER ANTAR KAS (tidak memengaruhi laba/rugi)")
        for t in data.transfers:
            label = f"{t.transfer_date.strftime('%d/%m/%Y')}  {t.from_account_name or '—'} → {t.to_account_name or '—'}"
            if t.description:
                label += f" ({t.description})"
            line(label, t.amount)

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 20
    return _xlsx_response(wb, f"LabaRugi_{date_from.strftime('%Y%m%d')}-{date_to.strftime('%Y%m%d')}.xlsx")
