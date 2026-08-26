"""Impor data member dari Excel (migrasi data lama).

Alur: unduh template → isi → unggah → PRATINJAU (dry-run, tak menyentuh DB) →
COMMIT (upsert per No. WA). Membuat akun + 1 "paket berjalan" (sisa sesi + expired).

Template kolom (baris pertama = header, persis):
  Nama | No. WA | Email | Kategori | Nama Paket | Sisa Sesi | Unlimited | Tanggal Expired | Tanggal Bergabung
"""
import io
from datetime import datetime, timezone, date, time

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.user import User, UserRole, MemberCategory
from app.models.package import MemberPackage, MemberPackageStatus
from app.core.security import get_password_hash
from app.services.whatsapp import normalize_phone
from app.services.quota import refresh_status

HEADERS = [
    "Nama", "No. WA", "Email", "Kategori", "Nama Paket",
    "Sisa Sesi", "Unlimited", "Tanggal Expired", "Tanggal Bergabung",
]

_CAT_MAP = {
    "bulanan": MemberCategory.BULANAN, "monthly": MemberCategory.BULANAN, "member": MemberCategory.BULANAN,
    "private": MemberCategory.PRIVATE, "privat": MemberCategory.PRIVATE, "pt": MemberCategory.PRIVATE,
    "per datang": MemberCategory.PER_DATANG, "per-datang": MemberCategory.PER_DATANG,
    "perdatang": MemberCategory.PER_DATANG, "per_datang": MemberCategory.PER_DATANG,
    "drop in": MemberCategory.PER_DATANG, "drop-in": MemberCategory.PER_DATANG, "dropin": MemberCategory.PER_DATANG,
}
_TRUE = {"ya", "yes", "y", "true", "1", "unlimited", "tak terbatas", "✓"}


def _s(v) -> str:
    return "" if v is None else str(v).strip()


def _parse_date(v):
    """Terima datetime/date atau string dd/mm/yyyy, dd-mm-yyyy, yyyy-mm-dd. Return date | None; raise ValueError kalau tak terbaca."""
    if v is None or _s(v) == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = _s(v)
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y", "%d %m %Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"tanggal tak terbaca: {s!r}")


def _parse_int(v):
    """Return int | None; raise ValueError kalau bukan angka."""
    if v is None or _s(v) == "":
        return None
    s = _s(v).replace(".", "").replace(",", "")
    if not s.lstrip("-").isdigit():
        raise ValueError(f"angka tak terbaca: {_s(v)!r}")
    return int(s)


def build_template_xlsx() -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Data Member"
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="BD7A61")
    for c, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = head_font
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center")
    # Contoh baris
    examples = [
        ["Sinta Dewi", "081234567890", "", "bulanan", "GOLD 1", 8, "", "30/09/2026", "01/03/2025"],
        ["Rara Putri", "0857-1111-2222", "rara@gmail.com", "private", "Private 10", 3, "", "15/10/2026", ""],
        ["Budi Santoso", "082199998888", "", "per_datang", "Tiket Drop-in", 1, "", "", ""],
        ["Maya Sari", "081300004444", "", "bulanan", "Unlimited", "", "ya", "31/08/2026", ""],
    ]
    for r, row in enumerate(examples, start=2):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)
    widths = [22, 18, 24, 12, 16, 10, 10, 16, 18]
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    info = wb.create_sheet("Petunjuk")
    lines = [
        ("PETUNJUK PENGISIAN", True),
        ("", False),
        ("• Nama & No. WA WAJIB diisi. No. WA jadi cara login member (harus unik).", False),
        ("• Email boleh dikosongkan — sistem membuat email otomatis (login tetap pakai No. WA).", False),
        ("• Kategori: bulanan / private / per_datang.", False),
        ("• Sisa Sesi: jumlah sesi tersisa saat ini. Kosongkan bila member tak punya paket aktif.", False),
        ("• Unlimited: isi 'ya' bila paket tak terbatas (Sisa Sesi diabaikan).", False),
        ("• Tanggal Expired & Bergabung: format dd/mm/yyyy (mis. 30/09/2026). Boleh kosong.", False),
        ("• Baris tanpa paket (Sisa Sesi kosong, Unlimited kosong, Expired kosong) → hanya dibuat akun.", False),
        ("• Hapus 4 baris contoh di sheet 'Data Member' sebelum mengunggah.", False),
        ("• Aman diulang: data dikenali per No. WA — menjalankan ulang tidak menggandakan.", False),
    ]
    for r, (txt, bold) in enumerate(lines, start=1):
        cell = info.cell(row=r, column=1, value=txt)
        if bold:
            cell.font = Font(bold=True, size=13, color="3B2E28")
    info.column_dimensions["A"].width = 95

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _parse_rows(file_bytes: bytes) -> list[dict]:
    """Baca xlsx → list baris mentah {row_no, raw fields}. Raise ValueError kalau header tak cocok."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb["Data Member"] if "Data Member" in wb.sheetnames else wb.worksheets[0]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        raise ValueError("File kosong.")
    header_norm = [_s(h).lower() for h in header]
    expected = [h.lower() for h in HEADERS]
    # Cocokkan berdasarkan urutan minimal Nama & No. WA di 2 kolom pertama
    if header_norm[:2] != expected[:2]:
        raise ValueError("Header tidak sesuai template. Unduh template dan gunakan kolom yang sama.")

    out = []
    for i, row in enumerate(rows_iter, start=2):
        if row is None or all(_s(c) == "" for c in row):
            continue  # lewati baris kosong
        vals = list(row) + [None] * (len(HEADERS) - len(row))
        out.append({
            "row_no": i,
            "nama": _s(vals[0]),
            "no_wa": _s(vals[1]),
            "email": _s(vals[2]),
            "kategori": _s(vals[3]),
            "nama_paket": _s(vals[4]),
            "sisa_sesi_raw": vals[5],
            "unlimited_raw": vals[6],
            "expired_raw": vals[7],
            "gabung_raw": vals[8],
        })
    return out


async def analyze(db: AsyncSession, file_bytes: bytes) -> dict:
    """Validasi + tentukan aksi (buat/perbarui) tanpa menyentuh DB."""
    parsed = _parse_rows(file_bytes)

    # Peta nomor → user existing
    existing = (await db.execute(select(User).where(User.phone.isnot(None)))).scalars().all()
    phone_to_user = {}
    for u in existing:
        n = normalize_phone(u.phone or "")
        if n:
            phone_to_user.setdefault(n, u)

    seen_phones: dict[str, int] = {}
    results = []
    n_create = n_update = n_with_pkg = n_error = 0

    for r in parsed:
        errors, warnings = [], []
        nama = r["nama"]
        if not nama:
            errors.append("Nama kosong")

        norm = normalize_phone(r["no_wa"])
        if not r["no_wa"]:
            errors.append("No. WA kosong")
        elif not norm:
            errors.append(f"No. WA tak valid: {r['no_wa']!r}")
        elif norm in seen_phones:
            errors.append(f"No. WA dobel dengan baris {seen_phones[norm]}")
        if norm and norm not in seen_phones:
            seen_phones[norm] = r["row_no"]

        # kategori
        cat = None
        if r["kategori"]:
            cat = _CAT_MAP.get(r["kategori"].lower())
            if cat is None:
                errors.append(f"Kategori tak dikenal: {r['kategori']!r} (pakai bulanan/private/per_datang)")
        else:
            cat = MemberCategory.PER_DATANG
            warnings.append("Kategori kosong → default per_datang")

        # angka & tanggal
        sisa = None
        unlimited = _s(r["unlimited_raw"]).lower() in _TRUE
        try:
            sisa = _parse_int(r["sisa_sesi_raw"])
            if sisa is not None and sisa < 0:
                errors.append("Sisa Sesi negatif")
        except ValueError as e:
            errors.append(f"Sisa Sesi: {e}")
        expired = gabung = None
        try:
            expired = _parse_date(r["expired_raw"])
        except ValueError as e:
            errors.append(str(e).capitalize())
        try:
            gabung = _parse_date(r["gabung_raw"])
        except ValueError:
            warnings.append("Tanggal Bergabung tak terbaca → diabaikan")

        has_pkg = unlimited or (sisa is not None) or (expired is not None) or bool(r["nama_paket"])
        action = "update" if (norm and norm in phone_to_user) else "create"

        if errors:
            n_error += 1
        elif action == "create":
            n_create += 1
        else:
            n_update += 1
        if has_pkg and not errors:
            n_with_pkg += 1

        results.append({
            "row_no": r["row_no"], "nama": nama, "no_wa": r["no_wa"], "norm_phone": norm,
            "email": r["email"] or None,
            "kategori": cat.value if cat else None,
            "nama_paket": r["nama_paket"] or ("Unlimited" if unlimited else ("Paket Migrasi" if has_pkg else "")),
            "sisa_sesi": sisa, "unlimited": unlimited,
            "expired": expired.isoformat() if expired else None,
            "gabung": gabung.isoformat() if gabung else None,
            "has_package": has_pkg, "action": action,
            "errors": errors, "warnings": warnings,
        })

    return {
        "total_rows": len(parsed),
        "to_create": n_create, "to_update": n_update, "with_package": n_with_pkg, "errors": n_error,
        "rows": results,
    }


async def commit(db: AsyncSession, file_bytes: bytes, default_password: str, actor_id=None) -> dict:
    """Jalankan upsert. Baris ber-error dilewati. Return ringkasan."""
    report = await analyze(db, file_bytes)
    created = updated = pkg_created = skipped = 0

    # Untuk email placeholder unik
    existing_emails = set(
        e.lower() for e in (await db.execute(select(User.email))).scalars().all() if e
    )

    def gen_email(norm_phone: str) -> str:
        base = f"{norm_phone}@reformeryourbody.com"
        cand = base
        i = 1
        while cand.lower() in existing_emails:
            cand = f"{norm_phone}-{i}@reformeryourbody.com"
            i += 1
        existing_emails.add(cand.lower())
        return cand

    now = datetime.now(timezone.utc)

    # Peta nomor → user existing (dibangun sekali; diperbarui saat membuat akun baru)
    all_users = (await db.execute(select(User).where(User.phone.isnot(None)))).scalars().all()
    phone_to_user = {}
    for u in all_users:
        n = normalize_phone(u.phone or "")
        if n:
            phone_to_user.setdefault(n, u)

    for row in report["rows"]:
        if row["errors"]:
            skipped += 1
            continue
        norm = row["norm_phone"]
        user = phone_to_user.get(norm)

        gabung = date.fromisoformat(row["gabung"]) if row["gabung"] else None
        # Simpan No. WA dalam format lokal ber-0 (mis. 62812… → 0812…) agar rapi & konsisten.
        phone_local = ("0" + row["norm_phone"][2:]) if (row["norm_phone"] or "").startswith("62") else (row["no_wa"] or None)
        if user is None:
            email = (row["email"] if row["email"] and "@" in row["email"] else None) or gen_email(norm)
            user = User(
                email=email, hashed_password=get_password_hash(default_password),
                full_name=row["nama"], phone=phone_local,
                role=UserRole.MEMBER, member_category=MemberCategory(row["kategori"]),
                join_date=gabung or now.date(), is_active=True,
            )
            db.add(user)
            await db.flush()
            phone_to_user[norm] = user
            created += 1
        else:
            user.full_name = row["nama"]
            user.member_category = MemberCategory(row["kategori"])
            if gabung:
                user.join_date = gabung
            updated += 1

        if row["has_package"]:
            # Hapus paket migrasi lama (package_id NULL & price 0 & ACTIVE) agar idempoten
            olds = (await db.execute(
                select(MemberPackage).where(
                    MemberPackage.member_id == user.id,
                    MemberPackage.package_id.is_(None),
                    MemberPackage.price_paid == 0,
                    MemberPackage.status == MemberPackageStatus.ACTIVE,
                )
            )).scalars().all()
            for o in olds:
                await db.delete(o)

            expired = date.fromisoformat(row["expired"]) if row["expired"] else None
            expires_at = datetime.combine(expired, time(23, 59), tzinfo=timezone.utc) if expired else None
            unlimited = row["unlimited"]
            sisa = None if unlimited else (row["sisa_sesi"] if row["sisa_sesi"] is not None else 0)
            mp = MemberPackage(
                member_id=user.id, package_id=None,
                package_name=row["nama_paket"] or "Paket Migrasi",
                is_unlimited=unlimited,
                monthly_expiry=(row["kategori"] == MemberCategory.BULANAN.value),
                sessions_total=sisa, sessions_remaining=sisa,
                price_paid=0, purchased_at=now, expires_at=expires_at,
                status=MemberPackageStatus.ACTIVE,
            )
            refresh_status(mp)
            db.add(mp)
            pkg_created += 1

    return {"created": created, "updated": updated, "packages": pkg_created, "skipped": skipped,
            "total_rows": report["total_rows"]}
